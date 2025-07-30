#!/usr/bin/env python3
"""
Test the flexible invoice number patterns.
"""

from enhanced_text_processor import EnhancedTextProcessor
from openpyxl import Workbook

def test_invoice_patterns():
    processor = EnhancedTextProcessor()
    wb = Workbook()
    ws = wb.active

    print("=" * 60)
    print("TESTING FLEXIBLE INVOICE NUMBER PATTERNS")
    print("=" * 60)
    
    # Test the invoice patterns you mentioned
    test_invoices = [
        'JF25003',      # letters + numbers
        'KB-20244',     # letters + dash + numbers  
        'MOTO123123E',  # letters + numbers + letter
        'MT-21233',     # letters + dash + numbers
        '12345',        # numeric only (legacy)
        'ABC123-456',   # general pattern
        'XY12',         # short but valid
        'A1',           # too short
        '123.45'        # decimal (should fail)
    ]
    
    label_match = {'category': 'invoice_label', 'replacement': 'JFINV'}

    print("Testing invoice number evaluation:")
    print("-" * 40)

    for invoice in test_invoices:
        ws.cell(row=1, column=1, value=invoice)
        cell = ws.cell(row=1, column=1)
        score = processor._evaluate_target_cell(cell, label_match, 10)
        result = '✅ SUCCESS' if score > 0 else '❌ FAILED'
        print(f"  {invoice:15s} -> Score: {score:2d} {result}")

    print("\nTesting invoice label detection:")
    print("-" * 40)
    
    test_labels = [
        'Invoice No:',
        'invoice no....',  
        'inv no###',
        'Invoice No: JF25003',
        'bill no!!!',
        'Invoice Number:',
        'INVOICE NO.',
    ]
    
    for label in test_labels:
        match = processor._find_label_match(label)
        if match:
            result = f"✅ MATCH -> {match['category']} -> {match['replacement']}"
        else:
            result = "❌ NO MATCH"
        print(f"  '{label:20s}' -> {result}")

    print("\n" + "=" * 60)
    print("Test completed!")

if __name__ == "__main__":
    test_invoice_patterns() 