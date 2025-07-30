#!/usr/bin/env python3
"""
Demo: Enhanced Text Replacement with Circular Pattern Checking

This script demonstrates the new enhanced text replacement feature that uses
circular pattern checking to find the correct target cell for replacement,
even when labels are positioned unusually.
"""

from openpyxl import Workbook
from enhanced_text_processor import EnhancedTextProcessor
import logging

# Set up logging to see the process
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


def create_demo_workbook():
    """Create a demo workbook with various label positioning scenarios."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Demo Sheet"
    
    # Scenario 1: Standard right positioning (D20 -> E20)
    ws['D20'] = 'Invoice No:'
    ws['E20'] = 'OLD-INV-123'
    
    # Scenario 2: Below positioning (D22 -> D23)
    ws['D22'] = 'Contract No:'
    ws['D23'] = 'OLD-CONTRACT-456'
    
    # Scenario 3: Above positioning (D25 -> D24)
    ws['D25'] = 'Reference No:'
    ws['D24'] = 'OLD-REF-789'
    
    # Scenario 4: Diagonal positioning (D27 -> E28)
    ws['D27'] = 'Date:'
    ws['E28'] = '2023-12-31'
    
    # Scenario 5: Far right positioning (D30 -> G30)
    ws['D30'] = 'ETD:'
    ws['G30'] = '2024-01-15'
    
    # Scenario 6: Left positioning (unusual but possible) (D32 -> C32)
    ws['D32'] = 'Dated:'
    ws['C32'] = '2023-11-30'
    
    # Scenario 7: Multiple potential targets (processor should pick best)
    ws['D35'] = 'Invoice Date:'
    ws['E35'] = 'HEADER'  # Less suitable (text)
    ws['F35'] = '2023-10-15'  # More suitable (date format)
    ws['G35'] = 'FOOTER'  # Less suitable (text)
    
    return wb


def demonstrate_enhanced_replacement():
    """Demonstrate the enhanced text replacement with circular pattern checking."""
    print("=" * 80)
    print("ENHANCED TEXT REPLACEMENT DEMO")
    print("=" * 80)
    print()
    
    # Create demo workbook
    print("1. Creating demo workbook with various label positioning scenarios...")
    wb = create_demo_workbook()
    ws = wb.active
    
    # Show original values
    print("\n2. Original values in the worksheet:")
    print("-" * 40)
    test_cells = [
        ('D20', 'E20'), ('D22', 'D23'), ('D25', 'D24'), 
        ('D27', 'E28'), ('D30', 'G30'), ('D32', 'C32'),
        ('D35', 'E35'), ('D35', 'F35'), ('D35', 'G35')
    ]
    
    for label_cell, value_cell in test_cells:
        label_val = ws[label_cell].value
        value_val = ws[value_cell].value
        if label_val or value_val:
            print(f"  {label_cell}: '{label_val}' -> {value_cell}: '{value_val}'")
    
    # Create enhanced processor
    print("\n3. Processing with Enhanced Text Processor...")
    print("-" * 40)
    processor = EnhancedTextProcessor()
    
    # Process the worksheet
    stats = processor.process_worksheet_with_circular_pattern(ws)
    
    # Show results
    print("\n4. Results after processing:")
    print("-" * 40)
    for label_cell, value_cell in test_cells:
        label_val = ws[label_cell].value
        value_val = ws[value_cell].value
        if label_val:
            print(f"  {label_cell}: '{label_val}' -> {value_cell}: '{value_val}'")
    
    # Show statistics
    print("\n5. Replacement Statistics:")
    print("-" * 40)
    total_replacements = sum(stats.values())
    print(f"  Total replacements: {total_replacements}")
    for category, count in stats.items():
        if count > 0:
            print(f"  {category}: {count} replacements")
    
    # Save demo file
    output_file = "enhanced_text_replacement_demo.xlsx"
    wb.save(output_file)
    print(f"\n6. Demo workbook saved as: {output_file}")
    
    return stats


def show_circular_pattern_explanation():
    """Show how the circular pattern checking works."""
    print("\n" + "=" * 80)
    print("CIRCULAR PATTERN CHECKING EXPLANATION")
    print("=" * 80)
    print()
    
    print("When a label is found at position D20, the processor checks adjacent cells")
    print("in this priority order:")
    print()
    
    pattern_explanation = [
        ("E20", "right", 10, "Most common position"),
        ("F20", "right+2", 8, "Extended right"),
        ("D21", "below", 7, "Below the label"),
        ("D19", "above", 6, "Above the label"),
        ("E21", "below-right", 5, "Diagonal below-right"),
        ("E19", "above-right", 5, "Diagonal above-right"),
        ("C20", "left", 4, "Left of label (unusual)"),
        ("C21", "below-left", 3, "Diagonal below-left"),
        ("C19", "above-left", 3, "Diagonal above-left"),
        ("G20", "right+3", 2, "Far right"),
        ("D22", "below+2", 2, "Far below"),
        ("D18", "above+2", 2, "Far above"),
    ]
    
    print("Priority | Position | Direction    | Description")
    print("-" * 60)
    for pos, direction, priority, description in pattern_explanation:
        print(f"{priority:8d} | {pos:8s} | {direction:12s} | {description}")
    
    print()
    print("The processor evaluates each candidate cell and selects the one with")
    print("the highest combined score (position priority + content suitability).")


def show_content_scoring_explanation():
    """Show how content scoring works for different categories."""
    print("\n" + "=" * 80)
    print("CONTENT SCORING EXPLANATION")
    print("=" * 80)
    print()
    
    print("The processor evaluates target cells based on their content:")
    print()
    
    scoring_rules = [
        ("Date/ETD patterns", [
            ("2024-01-15", "+20", "Strong date pattern match"),
            ("01/15/2024", "+15", "Alternative date format"),
            ("123456", "+5", "Contains numbers"),
            ("Some text", "+0", "No date indicators")
        ]),
        ("Code patterns (Invoice/Contract/Ref)", [
            ("ABC-DEF-123", "+20", "Strong code pattern"),
            ("INV/2024/001", "+15", "Code with separators"),
            ("ABC123", "+10", "Simple alphanumeric"),
            ("123456", "+5", "Numbers only"),
            ("Random text", "+0", "No code indicators")
        ]),
        ("General rules", [
            ("Short structured text", "+5", "Looks like data"),
            ("Very long text (>50 chars)", "-10", "Likely not a value"),
            ("Empty cell", "0", "Not suitable")
        ])
    ]
    
    for category, rules in scoring_rules:
        print(f"{category}:")
        for example, score, description in rules:
            print(f"  '{example}' -> {score:>3s} points ({description})")
        print()


if __name__ == "__main__":
    # Run the demonstration
    try:
        stats = demonstrate_enhanced_replacement()
        show_circular_pattern_explanation()
        show_content_scoring_explanation()
        
        print("\n" + "=" * 80)
        print("DEMO COMPLETE")
        print("=" * 80)
        print()
        print("Key Benefits of Enhanced Text Replacement:")
        print("✅ Handles unusual label positioning")
        print("✅ Works with foreign language labels")
        print("✅ Intelligent target cell selection")
        print("✅ Priority-based circular pattern checking")
        print("✅ Content-aware scoring system")
        print("✅ Flexible and extensible pattern matching")
        
    except Exception as e:
        print(f"Error during demonstration: {e}")
        import traceback
        traceback.print_exc() 