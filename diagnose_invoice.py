#!/usr/bin/env python3
"""
Diagnose invoice number detection in the user's Excel file.
"""

import sys
from enhanced_text_processor import EnhancedTextProcessor
from openpyxl import load_workbook

def diagnose_invoice(excel_file):
    print("=" * 60)
    print("DIAGNOSING INVOICE NUMBER DETECTION")
    print("=" * 60)
    print(f"File: {excel_file}")
    print()
    
    try:
        wb = load_workbook(excel_file)
        processor = EnhancedTextProcessor()
        
        for sheet_name in wb.sheetnames:
            print(f"🔍 SHEET: {sheet_name}")
            print("-" * 40)
            
            ws = wb[sheet_name]
            
            # Look for invoice-related cells
            invoice_cells = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        cell_str = str(cell.value).lower()
                        if any(keyword in cell_str for keyword in ['invoice', 'inv']):
                            invoice_cells.append((cell.coordinate, cell.value))
            
            if invoice_cells:
                print(f"Found {len(invoice_cells)} cells with 'invoice'/'inv':")
                for coord, value in invoice_cells:
                    print(f"  {coord}: '{value}'")
                    
                    # Test pattern matching
                    match = processor._find_label_match(value)
                    if match:
                        print(f"    ✅ MATCHES -> {match['category']} -> {match['replacement']}")
                        
                        # Test circular pattern search
                        actual_cell = ws[coord]
                        try:
                            target_cell = processor._find_target_cell_circular(ws, actual_cell, match)
                            if target_cell:
                                score = processor._evaluate_target_cell(target_cell, match, 10)
                                print(f"    🎯 Target: {target_cell.coordinate} = '{target_cell.value}' (score: {score})")
                            else:
                                print(f"    ❌ No target cell found")
                                
                                # Check adjacent cells manually
                                row_num = actual_cell.row
                                col_num = actual_cell.column
                                
                                adjacent_positions = [
                                    (row_num, col_num + 1, "right"),
                                    (row_num, col_num + 2, "right+2"),
                                    (row_num + 1, col_num, "below"),
                                ]
                                
                                print(f"    🔍 Adjacent cells:")
                                for r, c, direction in adjacent_positions:
                                    try:
                                        adj_cell = ws.cell(row=r, column=c)
                                        if adj_cell.value:
                                            score = processor._evaluate_target_cell(adj_cell, match, 10)
                                            print(f"      {direction}: {adj_cell.coordinate} = '{adj_cell.value}' (score: {score})")
                                    except:
                                        pass
                        except Exception as e:
                            print(f"    ❌ Error in search: {e}")
                    else:
                        print(f"    ❌ NO MATCH")
                print()
            else:
                print("No invoice-related cells found.")
                print()
                
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python diagnose_invoice.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    diagnose_invoice(excel_file) 