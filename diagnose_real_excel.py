#!/usr/bin/env python3
"""
Diagnose why date replacement fails on real Excel files.
"""

import sys
from enhanced_text_processor import EnhancedTextProcessor
from openpyxl import load_workbook

def diagnose_real_excel(excel_file):
    print("=" * 70)
    print("DIAGNOSING REAL EXCEL FILE")
    print("=" * 70)
    print(f"File: {excel_file}")
    print()
    
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        print(f"✅ Successfully loaded workbook")
        print(f"   Sheets: {wb.sheetnames}")
        print()
        
        # Create processor
        processor = EnhancedTextProcessor()
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"🔍 ANALYZING SHEET: {sheet_name}")
            print("-" * 50)
            
            ws = wb[sheet_name]
            
            # Look for date-related cells
            date_cells = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        # Check if it contains "date"
                        if "date" in str(cell.value).lower():
                            date_cells.append((cell.coordinate, cell.value))
            
            if date_cells:
                print(f"Found {len(date_cells)} cells containing 'date':")
                for coord, value in date_cells:
                    print(f"  {coord}: '{value}'")
                    
                    # Test label matching
                    match = processor._find_label_match(value)
                    if match:
                        print(f"    ✅ MATCHES -> {match['category']} -> {match['replacement']}")
                        
                        # Get the actual cell and test circular pattern
                        actual_cell = ws[coord]
                        try:
                            target_cell = processor._find_target_cell_circular(ws, actual_cell, match)
                            if target_cell:
                                score = processor._evaluate_target_cell(target_cell, match, 10)
                                print(f"    🎯 Target: {target_cell.coordinate} = '{target_cell.value}' (score: {score})")
                            else:
                                print(f"    ❌ No target cell found")
                                
                                # Check adjacent cells manually
                                row_num = actual_cell.row
                                col_num = actual_cell.column
                                
                                adjacent_positions = [
                                    (row_num, col_num + 1, "right"),
                                    (row_num, col_num + 2, "right+2"),
                                    (row_num + 1, col_num, "below"),
                                    (row_num - 1, col_num, "above")
                                ]
                                
                                print(f"    🔍 Checking adjacent cells:")
                                for r, c, direction in adjacent_positions:
                                    try:
                                        adj_cell = ws.cell(row=r, column=c)
                                        if adj_cell.value:
                                            score = processor._evaluate_target_cell(adj_cell, match, 10)
                                            print(f"      {direction}: {adj_cell.coordinate} = '{adj_cell.value}' (score: {score})")
                                    except:
                                        pass
                                        
                        except Exception as e:
                            print(f"    ❌ Error in circular search: {e}")
                            
                    else:
                        print(f"    ❌ NO MATCH")
                print()
            else:
                print("No cells containing 'date' found in this sheet.")
                print()
            
            # Check for table data area detection
            print("🛡️ CHECKING TABLE DATA AREA DETECTION:")
            table_cells_checked = 0
            for coord, value in date_cells[:3]:  # Check first 3 date cells
                cell = ws[coord]
                is_table_area = processor._is_likely_table_data_area(ws, cell)
                print(f"  {coord}: '{value}' -> {'TABLE AREA (SKIPPED)' if is_table_area else 'OK'}")
                table_cells_checked += 1
            
            if table_cells_checked == 0:
                print("  No date cells to check.")
            print()
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python diagnose_real_excel.py <excel_file>")
        print("Example: python diagnose_real_excel.py 'CT&INV&PL JF25035 FCA(1).xlsx'")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    diagnose_real_excel(excel_file) 