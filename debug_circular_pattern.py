#!/usr/bin/env python3
"""
Debug the circular pattern search to see why it's not finding adjacent date values.
"""

from enhanced_text_processor import EnhancedTextProcessor
from openpyxl import Workbook

def debug_circular_pattern():
    print("=" * 60)
    print("DEBUGGING CIRCULAR PATTERN SEARCH")
    print("=" * 60)
    
    # Create test workbook
    wb = Workbook()
    ws = wb.active
    
    # Set up test data
    ws.cell(row=1, column=1, value="Date:")         # A1
    ws.cell(row=1, column=2, value="25/07/2025")    # B1
    
    print("Test setup:")
    print("  A1: 'Date:'")
    print("  B1: '25/07/2025'")
    print()
    
    # Create processor
    processor = EnhancedTextProcessor()
    
    # Get the label cell (A1)
    label_cell = ws.cell(row=1, column=1)
    
    # Test label matching
    label_match = processor._find_label_match(label_cell.value)
    print(f"Label match for '{label_cell.value}':")
    print(f"  Result: {label_match}")
    print()
    
    if label_match:
        # Test circular pattern search
        print("Testing circular pattern search...")
        try:
            target_cell = processor._find_target_cell_circular(ws, label_cell, label_match)
            
            if target_cell:
                print(f"✅ Found target cell: {target_cell.coordinate}")
                print(f"   Target cell value: '{target_cell.value}'")
                
                # Test target cell evaluation
                score = processor._evaluate_target_cell(target_cell, label_match, 10)
                print(f"   Target cell score: {score}")
                
            else:
                print("❌ No target cell found!")
                
        except Exception as e:
            print(f"❌ Error in circular pattern search: {e}")
            
        # Let's manually check the adjacent cell regardless
        adjacent_cell = ws.cell(row=1, column=2)  # B1
        print(f"Manual check of B1 (adjacent cell):")
        print(f"  Value: '{adjacent_cell.value}'")
        try:
            score = processor._evaluate_target_cell(adjacent_cell, label_match, 10)
            print(f"  Score: {score}")
        except Exception as e:
            print(f"  Error evaluating: {e}")
    
    print("\n" + "=" * 60)

if __name__ == "__main__":
    try:
        debug_circular_pattern()
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc() 