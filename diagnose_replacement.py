#!/usr/bin/env python3
"""
Diagnose why date values aren't being replaced.
"""

from enhanced_text_processor import EnhancedTextProcessor
from openpyxl import Workbook
import logging

# Enable debug logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def diagnose_replacement():
    print("=" * 60)
    print("DIAGNOSING DATE REPLACEMENT PROCESS")
    print("=" * 60)
    
    # Create a simple test workbook
    wb = Workbook()
    ws = wb.active
    
    # Set up test data like your Excel might have
    ws.cell(row=1, column=1, value="Date:")
    ws.cell(row=1, column=2, value="25/07/2025")
    
    ws.cell(row=2, column=1, value="Invoice No:")  
    ws.cell(row=2, column=2, value="12345")
    
    ws.cell(row=3, column=1, value="Ref. No:")
    ws.cell(row=3, column=2, value="ABC123-456")
    
    print("Test data created:")
    print("  A1: 'Date:'       B1: '25/07/2025'")
    print("  A2: 'Invoice No:' B2: '12345'") 
    print("  A3: 'Ref. No:'    B3: 'ABC123-456'")
    print()
    
    # Create processor and process the worksheet
    processor = EnhancedTextProcessor()
    
    print("Starting replacement process...")
    print("-" * 40)
    
    # Process with detailed logging
    stats = processor.process_worksheet_with_circular_pattern(ws)
    
    print("-" * 40)
    print("Replacement process completed.")
    print()
    
    print("RESULTS:")
    print(f"Replacement statistics: {stats}")
    print()
    
    print("Final cell values:")
    print(f"  A1: '{ws.cell(row=1, column=1).value}' B1: '{ws.cell(row=1, column=2).value}'")
    print(f"  A2: '{ws.cell(row=2, column=1).value}' B2: '{ws.cell(row=2, column=2).value}'")
    print(f"  A3: '{ws.cell(row=3, column=1).value}' B3: '{ws.cell(row=3, column=2).value}'")
    print()
    
    # Check what changed
    expected_changes = [
        ("B1", "25/07/2025", "JFDATE"),
        ("B2", "12345", "JFINV"), 
        ("B3", "ABC123-456", "JFREF")
    ]
    
    print("ANALYSIS:")
    for cell_ref, original, expected in expected_changes:
        row, col = (1, 2) if cell_ref == "B1" else (2, 2) if cell_ref == "B2" else (3, 2)
        actual = ws.cell(row=row, column=col).value
        
        if actual == expected:
            print(f"  ✅ {cell_ref}: '{original}' -> '{actual}' (SUCCESS)")
        elif actual == original:
            print(f"  ❌ {cell_ref}: '{original}' -> NOT CHANGED (FAILED)")
        else:
            print(f"  ⚠️  {cell_ref}: '{original}' -> '{actual}' (UNEXPECTED)")

if __name__ == "__main__":
    try:
        diagnose_replacement()
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc() 